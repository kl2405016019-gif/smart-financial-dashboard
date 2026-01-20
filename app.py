import io
import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px
import openpyxl

st.set_page_config(page_title="Smart Financial Dashboard", layout="wide")

RATIO_ORDER = [
    "Gross Profit Margin",
    "Net Profit Margin",
    "ROA (Net Inc / Avg TA)",
    "ROE (Net Inc / Avg Equity)",
    "Current Ratio",
    "Quick Ratio",
    "Inventory Turnover",
    "Receivables Turnover",
    "Total Asset Turnover",
    "Collection Period (Days)",
    "Inventory Days",
    "Cash Turnover",
    "Working Capital Turnover",
    "PPE Turnover",
    "Debt to Equity",
    "Debt Ratio",
    "Long-term Debt to Equity",
    "Times Interest Earned",
    "P/E Ratio",
    "Earnings Yield",
    "Dividend Yield",
    "Dividend Payout Ratio",
    "Price-to-Book",
    "Altman Z-Score",
]

METRICS = [
    "Sales (Revenue)",
    "COGS",
    "EBIT",
    "Net Income",
    "Interest Expense",
    "Current Assets",
    "Inventory",
    "Accounts Receivable",
    "Current Liabilities",
    "Total Assets",
    "Total Liabilities",
    "Long-term Liabilities",
    "Total Equity (Book)",
    "Retained Earnings",
    "Cash & Cash Equivalents",
    "Marketable Securities",
    "PPE (Net)",
    "Shares Outstanding",
    "Dividends per Share",
    "Market Price per Share",
]

DEFAULT_YEARS = ["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]


def _safe_div(a, b):
    a = np.array(a, dtype=float)
    b = np.array(b, dtype=float)
    with np.errstate(divide="ignore", invalid="ignore"):
        out = np.where(b == 0, np.nan, a / b)
    return out


def _avg(series):
    # average with prior year; year1 uses itself
    s = np.array(series, dtype=float)
    prev = np.r_[s[0], s[:-1]]
    return (s + prev) / 2.0


def compute_ratios(df_inputs: pd.DataFrame) -> pd.DataFrame:
    """df_inputs index=Year labels; columns include METRICS."""
    g = df_inputs

    sales = g["Sales (Revenue)"].to_numpy(dtype=float)
    cogs = g["COGS"].to_numpy(dtype=float)
    ebit = g["EBIT"].to_numpy(dtype=float)
    ni = g["Net Income"].to_numpy(dtype=float)
    ie = g["Interest Expense"].to_numpy(dtype=float)

    ca = g["Current Assets"].to_numpy(dtype=float)
    inv = g["Inventory"].to_numpy(dtype=float)
    ar = g["Accounts Receivable"].to_numpy(dtype=float)
    cl = g["Current Liabilities"].to_numpy(dtype=float)

    ta = g["Total Assets"].to_numpy(dtype=float)
    tl = g["Total Liabilities"].to_numpy(dtype=float)
    ltd = g["Long-term Liabilities"].to_numpy(dtype=float)
    eq = g["Total Equity (Book)"].to_numpy(dtype=float)
    re = g["Retained Earnings"].to_numpy(dtype=float)

    cash = g["Cash & Cash Equivalents"].to_numpy(dtype=float)
    mktsec = g["Marketable Securities"].to_numpy(dtype=float)
    ppe = g["PPE (Net)"].to_numpy(dtype=float)

    sh = g["Shares Outstanding"].to_numpy(dtype=float)
    dps = g["Dividends per Share"].to_numpy(dtype=float)
    mps = g["Market Price per Share"].to_numpy(dtype=float)

    avg_ta = _avg(ta)
    avg_eq = _avg(eq)
    avg_inv = _avg(inv)
    avg_ar = _avg(ar)
    avg_cash = _avg(cash)
    avg_wc = _avg(ca - cl)
    avg_ppe = _avg(ppe)

    gp_margin = _safe_div(sales - cogs, sales)
    np_margin = _safe_div(ni, sales)
    roa = _safe_div(ni, avg_ta)
    roe = _safe_div(ni, avg_eq)
    current = _safe_div(ca, cl)
    quick = _safe_div(ca - inv, cl)

    inv_turn = _safe_div(cogs, avg_inv)
    recv_turn = _safe_div(sales, avg_ar)
    ta_turn = _safe_div(sales, avg_ta)
    coll_days = _safe_div(365.0, recv_turn)
    inv_days = _safe_div(365.0, inv_turn)

    cash_turn = _safe_div(sales, avg_cash)
    wc_turn = _safe_div(sales, avg_wc)
    ppe_turn = _safe_div(sales, avg_ppe)

    d2e = _safe_div(tl, eq)
    debt_ratio = _safe_div(tl, ta)
    ltd2e = _safe_div(ltd, eq)
    tie = _safe_div(ebit, ie)

    eps = _safe_div(ni, sh)
    pe = _safe_div(mps, eps)
    earn_yield = _safe_div(eps, mps)
    div_yield = _safe_div(dps, mps)
    payout = _safe_div(dps, eps)
    bvps = _safe_div(eq, sh)
    ptb = _safe_div(mps, bvps)

    wc = ca - cl
    mve = mps * sh
    z = (
        1.2 * _safe_div(wc, ta)
        + 1.4 * _safe_div(re, ta)
        + 3.3 * _safe_div(ebit, ta)
        + 0.6 * _safe_div(mve, tl)
        + 1.0 * _safe_div(sales, ta)
    )

    out = pd.DataFrame(
        {
            "Gross Profit Margin": gp_margin,
            "Net Profit Margin": np_margin,
            "ROA (Net Inc / Avg TA)": roa,
            "ROE (Net Inc / Avg Equity)": roe,
            "Current Ratio": current,
            "Quick Ratio": quick,
            "Inventory Turnover": inv_turn,
            "Receivables Turnover": recv_turn,
            "Total Asset Turnover": ta_turn,
            "Collection Period (Days)": coll_days,
            "Inventory Days": inv_days,
            "Cash Turnover": cash_turn,
            "Working Capital Turnover": wc_turn,
            "PPE Turnover": ppe_turn,
            "Debt to Equity": d2e,
            "Debt Ratio": debt_ratio,
            "Long-term Debt to Equity": ltd2e,
            "Times Interest Earned": tie,
            "P/E Ratio": pe,
            "Earnings Yield": earn_yield,
            "Dividend Yield": div_yield,
            "Dividend Payout Ratio": payout,
            "Price-to-Book": ptb,
            "Altman Z-Score": z,
        },
        index=g.index,
    )
    return out[RATIO_ORDER]


def load_from_excel(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    if "INPUT" not in wb.sheetnames:
        raise ValueError("Sheet 'INPUT' not found. Please upload the AC4313 template.")

    ws = wb["INPUT"]

    years = [ws.cell(r, 1).value for r in range(7, 12)]
    years = [y if y else DEFAULT_YEARS[i] for i, y in enumerate(years)]

    def read_company(start_col_year: int):
        # Header row (row 6): start_col_year is the column where 'Year' is
        headers = []
        for c in range(start_col_year, start_col_year + 21):
            headers.append(ws.cell(6, c).value)
        # normalize
        headers = [str(h).strip() if h else "" for h in headers]
        # rows 7-11 values
        data = []
        for r in range(7, 12):
            row = []
            for c in range(start_col_year, start_col_year + 21):
                v = ws.cell(r, c).value
                row.append(v)
            data.append(row)
        df = pd.DataFrame(data, columns=headers)
        df.insert(0, "Year", years)
        df = df.set_index("Year")
        # Keep only expected metrics where possible
        # Some columns might be missing/renamed; we'll align softly
        aligned = pd.DataFrame(index=df.index)
        for m in METRICS:
            if m in df.columns:
                aligned[m] = pd.to_numeric(df[m], errors="coerce")
            else:
                aligned[m] = np.nan
        return aligned

    comp_a = read_company(3)
    comp_b = read_company(25)

    # names from row 4 cols 3 and 25
    name_a = ws.cell(4, 3).value or "Company A"
    name_b = ws.cell(4, 25).value or "Company B"

    return str(name_a).strip(), str(name_b).strip(), comp_a, comp_b


def empty_inputs():
    df = pd.DataFrame(index=DEFAULT_YEARS, columns=METRICS, dtype=float)
    return df


def fmt_ratio_table(r: pd.DataFrame) -> pd.DataFrame:
    # Format: percentages for margins/yields; 2 dp for others
    percent_cols = {
        "Gross Profit Margin",
        "Net Profit Margin",
        "ROA (Net Inc / Avg TA)",
        "ROE (Net Inc / Avg Equity)",
        "Earnings Yield",
        "Dividend Yield",
    }
    out = r.copy()
    for c in out.columns:
        if c in percent_cols:
            out[c] = (out[c] * 100).round(2)
        else:
            out[c] = out[c].round(2)
    return out


def make_exec_prompts(r: pd.DataFrame) -> pd.DataFrame:
    # simple guided prompt table: Year1 vs Year5 change
    y1 = r.iloc[0]
    y5 = r.iloc[-1]
    delta = y5 - y1
    prompts = []
    for metric in [
        "Gross Profit Margin",
        "Net Profit Margin",
        "Current Ratio",
        "Debt Ratio",
        "Times Interest Earned",
        "Altman Z-Score",
    ]:
        direction = "increased" if delta[metric] > 0 else "decreased" if delta[metric] < 0 else "stayed flat"
        prompts.append(
            {
                "Metric": metric,
                "Year 1": y1[metric],
                "Year 5": y5[metric],
                "Δ Change": delta[metric],
                "Prompt": f"{metric} has {direction} from Year 1 to Year 5. Explain the likely business drivers and whether this is favourable for liquidity/profitability/risk.",
            }
        )
    return pd.DataFrame(prompts)


st.title("📊 Smart Financial Dashboard")
st.caption("Upload the AC4313 Excel template or start from blank inputs. Built for undergraduates & SMEs.")

with st.sidebar:
    st.header("Setup")
    uploaded = st.file_uploader("Upload AC4313 Excel (xlsx)", type=["xlsx"])
    use_sample = st.toggle("Use blank template", value=uploaded is None)
    st.divider()
    st.write("Display options")
    show_company_b = st.toggle("Show Company B", value=True)

if uploaded is not None and not use_sample:
    try:
        name_a, name_b, a_in, b_in = load_from_excel(uploaded.getvalue())
    except Exception as e:
        st.error(f"Could not read the file: {e}")
        st.stop()
else:
    name_a, name_b = "Company A", "Company B"
    a_in, b_in = empty_inputs(), empty_inputs()

# Editable inputs
st.subheader("1) Data Inputs")
col1, col2 = st.columns([1, 1], gap="large")
with col1:
    st.markdown(f"### 🅰️ {name_a}")
    a_edit = st.data_editor(
        a_in,
        use_container_width=True,
        num_rows="fixed",
        key="a_inputs",
    )
with col2:
    st.markdown(f"### 🅱️ {name_b}")
    if show_company_b:
        b_edit = st.data_editor(
            b_in,
            use_container_width=True,
            num_rows="fixed",
            key="b_inputs",
        )
    else:
        st.info("Company B is hidden (toggle in sidebar).")
        b_edit = b_in.copy()

# Compute
try:
    a_rat = compute_ratios(a_edit)
    b_rat = compute_ratios(b_edit) if show_company_b else None
except Exception as e:
    st.error(f"Could not compute ratios. Please check inputs for missing or invalid values. Details: {e}")
    st.stop()

# KPIs
st.subheader("2) Key KPIs")

def kpi_block(title, r_df):
    last = r_df.iloc[-1]
    first = r_df.iloc[0]
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Net Profit Margin (%)", f"{(last['Net Profit Margin']*100):.2f}", f"{((last['Net Profit Margin']-first['Net Profit Margin'])*100):.2f}")
    c2.metric("ROA (%)", f"{(last['ROA (Net Inc / Avg TA)']*100):.2f}")
    c3.metric("Current Ratio", f"{last['Current Ratio']:.2f}")
    c4.metric("Debt Ratio", f"{last['Debt Ratio']:.2f}")
    c5.metric("Altman Z", f"{last['Altman Z-Score']:.2f}")

k1, k2 = st.columns([1, 1], gap="large")
with k1:
    st.markdown(f"#### 🅰️ {name_a}")
    kpi_block(name_a, a_rat)
with k2:
    if show_company_b:
        st.markdown(f"#### 🅱️ {name_b}")
        kpi_block(name_b, b_rat)

# Ratio tables
st.subheader("3) Ratio Tables")
tab1, tab2 = st.tabs(["Company A", "Company B"])
with tab1:
    st.dataframe(fmt_ratio_table(a_rat), use_container_width=True)
with tab2:
    if show_company_b:
        st.dataframe(fmt_ratio_table(b_rat), use_container_width=True)
    else:
        st.info("Company B is hidden.")

# Charts
st.subheader("4) Trend Charts")
all_metrics = RATIO_ORDER
metric = st.selectbox("Choose a ratio to plot", all_metrics, index=0)

plot_df = pd.DataFrame({"Year": a_rat.index, name_a: a_rat[metric].values})
if show_company_b:
    plot_df[name_b] = b_rat[metric].values
plot_long = plot_df.melt(id_vars="Year", var_name="Company", value_name="Value")
fig = px.line(plot_long, x="Year", y="Value", color="Company", markers=True)
st.plotly_chart(fig, use_container_width=True)

# Exec summary helper
st.subheader("5) Executive Summary Helper")
colA, colB = st.columns(2, gap="large")
with colA:
    st.markdown(f"##### 🅰️ {name_a}")
    st.dataframe(make_exec_prompts(a_rat), use_container_width=True)
with colB:
    if show_company_b:
        st.markdown(f"##### 🅱️ {name_b}")
        st.dataframe(make_exec_prompts(b_rat), use_container_width=True)

# Downloads
st.subheader("6) Download")

csv_a = a_rat.reset_index().to_csv(index=False).encode("utf-8")
st.download_button("Download Company A ratios (CSV)", data=csv_a, file_name="company_a_ratios.csv", mime="text/csv")

if show_company_b:
    csv_b = b_rat.reset_index().to_csv(index=False).encode("utf-8")
    st.download_button("Download Company B ratios (CSV)", data=csv_b, file_name="company_b_ratios.csv", mime="text/csv")

# Excel report
output = io.BytesIO()
with pd.ExcelWriter(output, engine="openpyxl") as writer:
    a_edit.to_excel(writer, sheet_name="Inputs_CompanyA")
    a_rat.to_excel(writer, sheet_name="Ratios_CompanyA")
    if show_company_b:
        b_edit.to_excel(writer, sheet_name="Inputs_CompanyB")
        b_rat.to_excel(writer, sheet_name="Ratios_CompanyB")
    # Combined for easy teaching
    comb = a_rat.copy()
    comb.columns = pd.MultiIndex.from_product([[name_a], comb.columns])
    if show_company_b:
        comb2 = b_rat.copy()
        comb2.columns = pd.MultiIndex.from_product([[name_b], comb2.columns])
        combined = pd.concat([comb, comb2], axis=1)
    else:
        combined = comb
    combined.to_excel(writer, sheet_name="Combined")

st.download_button(
    "Download Excel report",
    data=output.getvalue(),
    file_name="smart_financial_dashboard_report.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

st.caption("Tip: For teaching, ask students to justify ratio changes using notes to accounts + business events.")
